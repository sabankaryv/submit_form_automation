import os
import pandas as pd
import numpy as np
import openpyxl
import requests
import playwright
from playwright.sync_api import sync_playwright,expect
import openpyxl
import requests
import re
import time
import random
from subprocess import *
import threading
from bs4 import BeautifulSoup
script_dir = os.path.dirname(os.path.abspath(__file__))
import os
os.environ['PLAYWRIGHT_BROWSERS_PATH'] = 'chromium'

unameList = ['uname' , 'user_name','userName',"name"]
lastName=['lastName','lname','last_name']
subject=['subject','sub']
captchalist=['captcha_code']
content=['content']
userData = {"name":"guest",'email':'guest@gmail.com'}
code='9*9*9*9*'

def playwrt(url,first_name, last_name, subject, comment, email):
    with sync_playwright() as p:
        
        print("-"*50)
        print("Processing For Domain {}".format(url))
        print("-"*50)
        browser = p.chromium.launch(headless=False)
        #browser = p.firefox.launch(headless=False)        
        context = browser.new_context(viewport={"width": 1920, "height": 1080})
        page = context.new_page()
        page.goto(url)
        contact_link = page.query_selector("a[href*='contact']")
        if contact_link:
            contact_url = contact_link.get_attribute("href")
            new_browser = p.chromium.launch(headless=False)
            #new_browser = p.firefox.launch(headless=False)
            context = browser.new_context(viewport={"width": 1920, "height": 1080})
            new_page = context.new_page()
            new_page.goto(contact_url,wait_until="load")
            #print("Contact Page {}".format(contact_url))
            iframe_element = new_page.wait_for_selector('iframe')
            if iframe_element:
                iframe_src = iframe_element.get_attribute('src')
                file_name = iframe_src.split('/')[-1]
                #print(file_name)
                form_new_browser = p.chromium.launch(headless=False)
                form_context = browser.new_context(viewport={"width": 1920, "height": 1080})
                form_page = form_context.new_page()
                form_page.goto(url+str("/"+file_name), wait_until="load")
                time.sleep(0.2)
                html_content = form_page.content()
                soup = BeautifulSoup(html_content, "html.parser")
                #print(soup)
                form = soup.find("form")
                form_elements = form.find_all(["input","textarea"])
                #print(form_elements)
                for element in form_elements:
                    key = element.get("type")
                    value = element.get("name")
                    my_dict = {key: value}
                    for key, value in my_dict.items():
                            if key == 'text':
                                value = my_dict.get(key)
                                #print(value)
                                if(value in unameList):
                                    enterNameValue=form_page.locator("(//input[(@name='"+value+"')])")
                                    for char in first_name:
                                        enterNameValue.type(char)
                                        time.sleep(0.20)
                                if(value in subject):
                                    enterNameValue=form_page.locator("(//input[(@name='"+value+"')])")
                                    for char in subject:
                                        enterNameValue.type(char)
                                        time.sleep(0.20)
                                if(value in captchalist):
                                    enterNameValue=form_page.locator("(//input[(@name='"+value+"')])")
                                    for char in code:
                                        enterNameValue.type(char)
                                        time.sleep(0.20)
                                    # time.sleep(3)
                                    # enterNameValue.type("9*9*9*9*")
                                    # time.sleep(5)
                            if key == 'email':
                                value = my_dict.get(key)
                                if value!=None:
                                    #print(value,type(str(value)))
                                    enterNameValue=form_page.locator("(//input[(@name='"+value+"')])")
                                    input_value = email
                                    for char in input_value:
                                        enterNameValue.type(char)
                                        time.sleep(0.30)
                            if key is None:
                                value = my_dict.get(key)
                                if(value in content):
                                    enterNameValue=form_page.locator("textarea")
                                    time.sleep(3)
                                    enterNameValue.type(comment)
                            if key == 'submit':
                                #print("key is --- ",key,"\n")
                                time.sleep(10)
                                submitLocator = form_page.locator("input[type='submit']")
                                if submitLocator!="":
                                    #print("This is submit")
                                    time.sleep(30) 
                                    submitLocator.click()
                                    print("Success")
                            # else:
                            #     print("No submit ")    
        else:
            print("Contact page not found.")

def process_excel_data(Domain_file,Data_file):
    if mode==1:
        #print("Mode 1")
        wb1 = openpyxl.load_workbook(Domain_file)
        wb2 = openpyxl.load_workbook(Data_file)
        sheet1 = wb1.active
        sheet2 = wb2.active
        for row1 in sheet1.iter_rows(values_only=True):
            domain = row1[0]
            for row2 in sheet2.iter_rows(values_only=True):
                first_name = row2[0]    
                last_name=row2[1]
                subject = row2[2]
                comment=row2[3]
                email = row2[4]
                #print(f"fname: {first_name},{last_name},{subject},{comment},{email} domain: {domain}")
                playwrt(domain,first_name, last_name, subject, comment, email)
    else:
        #print("Mode 2")
        wb1 = openpyxl.load_workbook(Domain_file)
        wb2 = openpyxl.load_workbook(Data_file)
        sheet1 = wb1.active
        sheet2 = wb2.active
        max_rows_domain = sheet1.max_row
        max_rows_fname = sheet2.max_row
        for i in range(1, min(max_rows_domain, max_rows_fname) + 1):
            domain = sheet1.cell(row=i, column=1).value
            first_name = sheet2.cell(row=i, column=1).value
            last_name = sheet2.cell(row=i, column=2).value
            subject = sheet2.cell(row=i, column=3).value
            comment = sheet2.cell(row=i, column=4).value
            email = sheet2.cell(row=i, column=5).value
            playwrt(domain,first_name, last_name, subject, comment, email)
            #print(f": {first_name},{last_name},{subject},{comment},{email} domain: {domain}")




print("*"*50)
mode = int(input("Choose below option To Start Process \n 1]One To One \n 2]One Two Many \n : "))
if (mode != 1 and mode != 2):
    print("Please enter valid input")
    exit()
Domain_file = input("Enter the Domain  file name: ")
Data_file= input("Enter the Data  file name: ")
print("*"*50)
process_excel_data(Domain_file,Data_file)